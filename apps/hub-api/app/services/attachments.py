"""Attachment text extraction — PDF, DOCX, XLSX, and vision model for scans/images."""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass

import httpx

from app.core.config import get_settings

logger = logging.getLogger(__name__)

MAX_EXTRACT_CHARS = 8000
MAX_VISION_BYTES = 10 * 1024 * 1024  # 10 MB — guard before base64 encoding
SPARSE_TEXT_THRESHOLD = 50  # fewer chars from PDF → treat as scan


@dataclass
class AttachmentContent:
    filename: str
    mime_type: str
    text: str
    page_count: int | None = None
    is_scan: bool = False
    extraction_method: str = "text"


def extract_text(filename: str, data: bytes) -> AttachmentContent | None:
    """Route to the appropriate extractor based on file extension."""
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return _extract_pdf(filename, data)
    elif lower.endswith(".docx"):
        return _extract_docx(filename, data)
    elif lower.endswith(".xlsx"):
        return _extract_xlsx(filename, data)
    elif lower.endswith(".xls"):
        logger.info("attachment_legacy_xls filename=%s", filename)
        return AttachmentContent(
            filename=filename,
            mime_type="application/vnd.ms-excel",
            text="",
            extraction_method="unsupported_legacy_format",
        )
    elif lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
        return _extract_image(filename, data)
    elif lower.endswith((".txt", ".csv")):
        return _extract_plaintext(filename, data)
    else:
        logger.info("attachment_unsupported filename=%s", filename)
        return None


def _extract_pdf(filename: str, data: bytes) -> AttachmentContent:
    """Extract text from PDF using pymupdf4llm if available, fallback to pymupdf."""
    text = ""
    page_count = None
    try:
        import pymupdf

        doc = pymupdf.open(stream=data, filetype="pdf")
        page_count = len(doc)
        pages = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        text = "\n\n".join(pages)
    except ImportError:
        logger.info("pymupdf_not_installed filename=%s", filename)
        text = ""
    except Exception as exc:
        logger.info("pdf_extract_failed filename=%s err=%s", filename, type(exc).__name__)
        text = ""

    # Check if this is a scanned PDF (very little extractable text)
    is_scan = len(text.strip()) < SPARSE_TEXT_THRESHOLD and len(data) > 10000
    if is_scan:
        vision_text = _call_vision_model(filename, data)
        if vision_text:
            return AttachmentContent(
                filename=filename,
                mime_type="application/pdf",
                text=vision_text[:MAX_EXTRACT_CHARS],
                page_count=page_count,
                is_scan=True,
                extraction_method="vision",
            )

    return AttachmentContent(
        filename=filename,
        mime_type="application/pdf",
        text=text[:MAX_EXTRACT_CHARS],
        page_count=page_count,
        is_scan=is_scan,
        extraction_method="text",
    )


def _extract_docx(filename: str, data: bytes) -> AttachmentContent:
    """Extract text from DOCX using python-docx."""
    text = ""
    try:
        from docx import Document

        doc = Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
    except ImportError:
        logger.info("python_docx_not_installed filename=%s", filename)
    except Exception as exc:
        logger.info("docx_extract_failed filename=%s err=%s", filename, type(exc).__name__)

    return AttachmentContent(
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        text=text[:MAX_EXTRACT_CHARS],
        extraction_method="text",
    )


def _extract_xlsx(filename: str, data: bytes) -> AttachmentContent:
    """Extract text from XLSX (first sheet summary) using openpyxl."""
    text = ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is not None:
                rows = []
                for row in ws.iter_rows(max_row=100, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(" | ".join(cells))
                text = "\n".join(rows)
        finally:
            wb.close()
    except ImportError:
        logger.info("openpyxl_not_installed filename=%s", filename)
    except Exception as exc:
        logger.info("xlsx_extract_failed filename=%s err=%s", filename, type(exc).__name__)

    return AttachmentContent(
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        text=text[:MAX_EXTRACT_CHARS],
        extraction_method="text",
    )


def _extract_image(filename: str, data: bytes) -> AttachmentContent:
    """Route images to vision model for description/OCR."""
    vision_text = _call_vision_model(filename, data)
    return AttachmentContent(
        filename=filename,
        mime_type=_guess_image_mime(filename),
        text=vision_text[:MAX_EXTRACT_CHARS] if vision_text else "",
        is_scan=True,
        extraction_method="vision" if vision_text else "none",
    )


def _extract_plaintext(filename: str, data: bytes) -> AttachmentContent:
    """Plain text / CSV — decode directly."""
    try:
        text = data.decode("utf-8", errors="replace")
    except Exception:
        text = ""
    return AttachmentContent(
        filename=filename,
        mime_type="text/plain",
        text=text[:MAX_EXTRACT_CHARS],
        extraction_method="text",
    )


def _call_vision_model(filename: str, data: bytes) -> str | None:
    """Call the vLLM vision model endpoint for image/scan analysis."""
    if len(data) > MAX_VISION_BYTES:
        logger.info("vision_skipped_too_large filename=%s size=%d", filename, len(data))
        return None
    try:
        settings = get_settings()
        if settings.inference_mode.lower() != "vllm":
            logger.info("vision_skipped_not_vllm filename=%s", filename)
            return None

        import base64

        mime = _guess_image_mime(filename)
        b64 = base64.b64encode(data).decode("ascii")

        # Vision model on port 8004 (separate vLLM service)
        vision_url = getattr(settings, "vllm_vision_url", "http://localhost:8004/v1")
        vision_model = getattr(settings, "vllm_vision_model", "Qwen/Qwen2.5-VL-7B")

        with httpx.Client(timeout=60.0) as client:
            resp = client.post(
                f"{vision_url}/chat/completions",
                json={
                    "model": vision_model,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "image_url",
                                    "image_url": {"url": f"data:{mime};base64,{b64}"},
                                },
                                {
                                    "type": "text",
                                    "text": (
                                        "Describe this document/image content for an email assistant. "
                                        "Extract all visible text, tables, and key information. "
                                        "Be concise but complete. If it's a form or invoice, extract the key fields."
                                    ),
                                },
                            ],
                        }
                    ],
                    "max_tokens": 512,
                    "temperature": 0.1,
                },
            )
            if resp.status_code >= 400:
                logger.info("vision_model_http_error status=%s filename=%s", resp.status_code, filename)
                return None
            result = resp.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            return content.strip() or None
    except httpx.TimeoutException:
        logger.info("vision_model_timeout filename=%s", filename)
        return None
    except Exception as exc:
        logger.info("vision_model_failed filename=%s err=%s", filename, type(exc).__name__)
        return None


def _guess_image_mime(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".png"):
        return "image/png"
    elif lower.endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    elif lower.endswith(".gif"):
        return "image/gif"
    elif lower.endswith(".webp"):
        return "image/webp"
    elif lower.endswith(".pdf"):
        return "application/pdf"
    return "application/octet-stream"


def summarize_attachment(text: str, filename: str) -> str:
    """Generate a 1-2 sentence summary of extracted attachment content using 27B classify model."""
    if not text or not text.strip():
        return f"Attachment '{filename}' could not be analyzed."
    try:
        settings = get_settings()
        if settings.inference_mode.lower() != "vllm":
            # Stub/Ollama: return first 200 chars as summary
            snippet = text[:200].replace("\n", " ").strip()
            return f"{filename}: {snippet}..."

        with httpx.Client(timeout=10.0) as client:
            resp = client.post(
                f"{settings.vllm_classify_url.rstrip('/')}/chat/completions",
                json={
                    "model": settings.vllm_classify_model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "Summarize the following attachment content in 1-2 sentences for an email assistant. Be factual and concise.",
                        },
                        {
                            "role": "user",
                            "content": f"Filename: {filename}\n\nContent:\n{text[:2000]}",
                        },
                    ],
                    "max_tokens": 80,
                    "temperature": 0.1,
                },
            )
            if resp.status_code >= 400:
                snippet = text[:150].replace("\n", " ").strip()
                return f"{filename}: {snippet}..."
            result = resp.json()
            summary = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            return summary.strip() or f"{filename}: content extracted."
    except Exception:
        snippet = text[:150].replace("\n", " ").strip()
        return f"{filename}: {snippet}..."
